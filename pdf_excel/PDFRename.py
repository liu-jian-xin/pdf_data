# encoding:utf-8


import time
import os
import sys

import pdfplumber
import re
import tabula
import numpy as np
from openpyxl import workbook
import logging



#文件重命名
def rename(olddir,dirname,new_name):
    try:
        new_name_string = []
        # 如果名字中含有/，用-替代，因为pdf名字中不能有/
        for i in new_name:
            new_name_string.append(''.join(i).replace('/','-'))
        new_name_str = '_'.join(new_name_string)
        filetype = os.path.splitext(olddir)[1]
        newdir = os.path.join(dirname,new_name_str + filetype)
        os.rename(olddir, newdir)
    except FileExistsError as e:
        print(e)


# 加载配置文件
def loadSettingFile(KEYWORDS_Path):
    logging.info('>>>Loading setting file:%s' % os.path.basename(KEYWORDS_Path))
    PathList = {}  # 储存路径列表
    with open(KEYWORDS_Path, 'r', encoding='UTF-8') as fp:
        lines_kw = fp.readlines( )
        for line in lines_kw:
            # print(line)
            line = line.rstrip('\n')  # 删除行尾的换行符
            if re.match(r'^#', line):  # 注释内容，忽略
                pass
            else:
                Type, Path = line.split('=')  # 获得路径
                PathList[Type] = Path
                logging.info('>>>Content:\n %s' % PathList)
    logging.info('>>>Loading setting file done!')
    return PathList


# Marketing表解析规则
def pdf_marketing(text):
    # Marketing表有两种格式，分别用两种不同的匹配规则
    rules = ['Ungebundener (.*?) - '
        , 'Belegnummer: (.*?)\n'
        , 'Hamburg, (.*?)\n'
        , '\n(.*) - \('
        , '\(Basis.*\n(.*)\)'
        , 'Währung.*\n(.*)\n'
        , '\n(.*)\nRechnungsbetrag'
        , ' (\S*?) Stück'
        , 'Stück (.*?) '
        , 'Stück .* (.*?)EUR'
        , 'Umsatzsteuer \((.*?)\)'
        , 'Umsatzsteuer.*\) (.*?)EUR'
        , 'Gesamt Rechnungsbetrag.*\) (.*?)EUR']

    if "".join(re.findall(rules[1], text)) == '':
        rules = ['Gebundener (.*?) -'  # Gebundener (.*?) -
            , 'Rechnungsnummer: (.*?)\n'  # Rechnungsnummer: (.*?)\n
            , 'Hamburg, (.*?)\n'
            , '\n(.*) - Leistungszeitraum'  # \n(.*) - Leistungszeitraum
            , 'Leistungszeitraum (.*?)\n'  # Leistungszeitraum (.*?)\n
            , 'Währung.*\n(.*)\n'
            , '\n(.*)\nRechnungsbetrag'
            , ' (\S*?) Stück'
            , 'Stück (.*?) '
            , 'Stück .* (.*?)EUR'
            , 'Umsatzsteuer \((.*?)\)'
            , 'Umsatzsteuer.*\) (.*?)EUR'
            , 'Gesamt Rechnungsbetrag.*\) (.*?)EUR']
    str1="".join(re.findall(rules[5],text))
    if str1.find('EUR')>0:
        rules[5]='Währung.*\n[0-9]* (.*?) [0-9]'
    str2="".join(re.findall(rules[6],text))
    if str2.find('EUR') > 0:
        rules[6]='abcdefgh'
    words = []
    new_name = ['OTTO_Marekting']
    for index in range(len(rules)):
        rule = rules[index]
        word = re.findall(rule,text)
        if index >= 7:
            word = word[0].replace('.','').replace(',','.')
        if index == 1:
            new_name.append(''.join(word))
        if index == 2:
            word_list = ''.join(word).replace('.', '').replace('Januar', '01').replace('Februar', '02').replace('März', '03').replace('April', '04').replace('Mai', '05').replace('Juni', '06').replace('Juli', '07').replace('August','08').replace('September', '09').replace('Oktober', '10').replace('November', '11').replace('Dezember', '12').split(' ')
            new_name.append(''.join(word_list[::-1]))
            words.append('.'.join(word_list[::-1]))
            continue
        words.append(word)

    for index in range(len(words)):
        if index >= 7:
            if words[index].find('%') >= 0:
                words[index] = float(words[index].replace('%',''))/100
            else:
                words[index] = float(words[index])
    return words,new_name


#  Rework表名字
def pdf_Rework_name(text):
    new_name = ['OTTO_Rework']
    rules1 = ['RECHNUNG - BELASTUNG (.*)\n'
        ,'Leistungsdatum:\s*(.*?)\n']
    for index in range(len(rules1)):
        rule1 = rules1[index]
        word = re.findall(rule1, text)
        if index == 0:
            word = word[0].replace(' ','')
        if index == 1 :
            word = ''.join(list(reversed(word[0].split('.'))))
        new_name.append(word)
    return new_name


# Rework表解析 26
def pdf_Rework(text):
    rules =['RECHNUNG - (.*)\n'
        ,'KONTROLLBERICHTSNUMMER\n.* ([0-9]*/[0-9]*)\n'
        ,'LIEFERMENGE\n([0-9]*)'
        ,'LIEFERMENGE\n[0-9]* (.*?) '
        ,'LIEFERMENGE\n[0-9]* .*? ([0-9]*) '
        ,'LIEFERMENGE\n.* ([0-9]*) Stk./Set'
        ,'STYLE NR\.\n([0-9]*)\s*'
        ,'STYLE NR\.\n[0-9]* ([0-9]*)\s*'
        ,'STYLE NR\.\n[0-9]* [0-9]* ([0-9]*)\s*'
        ,'STYLE NR\.\n.* (.*?)\n'
        ,'ARTIKELBESCHREIBUNG\n(.*)\n'
        ,'REKLAMATIONSGRUND\n([\s\S]*)\nNACHBEARBEITUNGSANWEISUNG'
        ,'NACHBEARBEITUNGSANWEISUNG\n([\s\S]*)\nPREISNACHLASS'
        ,'Leistungsdatum:\s*(.*?)\n'
        ,'NACHBEARBEITETE MENGE\s*([0-9]*)'
        ,'ohne Fehler\s*([0-9]*)'
        ,'korrigiert \(lagerfähig\)\s*([0-9]*)'
        ,'abgelehnt \(nicht lagerfähig\)\s*([0-9]*)'
        ,'ORIGINAL KOSTEN\s*(.*?) '
        ,'Nachbearbeitungskosten / Sonstiges.* (.*?) EUR'
        ,'Warenwert\s*\(Kostenart 2135200200\).* (.*?) EUR'
        ,'Reduktionsbetrag\s*(.*?) EUR'
        ,'Claim Kosten:\s*(.*?) EUR'
        ,'MwSt.\s*([0-9]* \%)\s*'
        ,'MwSt.*\%:\s*(.*?) EUR'
        ,'Gesamt:\s*(.*?) EUR'
           ]
    words = []
    for index in range(0,len(rules)):

        rule = rules[index]
        word = re.findall(rule,text)
        if index == 13:
            word = '.'.join(list(reversed(word[0].split('.'))))
        if index >= 18 :
            word = word[0].replace('.','').replace(',','.')
        words.append(word)
    for index in range(len(words)):
        if index >= 18:
            if words[index].find('%') >= 0:
                words[index] = float(words[index].replace('%','').strip())/100
            else:
                words[index] = float(words[index])
    return words


# 年返表解析 13
def pdf_year(text):
    rules =['Rechnungsnummer:\n([0-9]*?)\n'
        ,'Belegdatum:\n.* (.*?)\n'
        ,'\n(Abrechnungszeitraum.*?)\n'
        ,'Umsatzanteil \(HAWA\).*\n.* (.*?) EUR'
        ,'Konditionenart.*\n(.*?) [0-9]'
        ,'Konditionenart.*\n.* (.*?) \%'
        ,'Konditionenart.*\n.*EUR\s*(.*) EUR.*?EUR.*?EUR'
        ,'Konditionenart.*\n.*EUR\s*(.*) EUR.*?EUR'
        ,'Konditionenart.*\n.*EUR(.*)EUR\n'
        ,'Gesamt Netto Zahlbetrag.*\nUST (.*\%) '
        ,'Gesamt Netto Zahlbetrag.*\n.*\%.* (.*?) EUR.* EUR'
        ,'Gesamt Netto Zahlbetrag.*\n.* EUR(.*?) EUR'
        ,'Gesamt Brutto Zahlbetrag\*(.*?) EUR\n'
        ,'Gesamt Netto Zahlbetrag.*\n.*\nUST (.*\%) '
        ,'Gesamt Netto Zahlbetrag.*\n.*\n.*\%.* (.*?) EUR.* EUR\n'
        ,'Gesamt Netto Zahlbetrag.*\n.*\n.*\%.* (.*?) EUR\n'
           ]
    words = []
    new_name = ['OTTO_年返']

    for index in range(len(rules)):

        rule = rules[index]
        word = re.findall(rule,text)
        if index == 0:
            new_name.append(word)
        if index == 1:
            # print(word[0])
            word = '.'.join(list(reversed(word[0].split('.'))))
            word_list = word.replace('.','')
            new_name.append(word_list)
        if index == 6:
            word = word[0]+ ' EUR'
        if index >= 9:
            word = word[0].replace('.','').replace(',','.')
        words.append(word)
    for index in range(len(words)):
        if index == 11 or index == 13 or index == 15 or index == 9:
            if words[index].find('%') >= 0:
                # print(words[index])
                words[index] = float(words[index].replace('%','').strip())/100
            else:
                # print(words[index])
                words[index] = float(words[index].strip())
        # print(words)
    return words,new_name


# 量差表解析规则
def pdf_amount(text):
    rules = ['.*(B\w* M\w*)\n'
        , 'Unsere Referenz.*\n.*\n.* (.*?)\n'
        , 'Datum\n.*-\S*\s(.*?)\n'
        , 'Ihre Rechnung.*\n(\d*?) '
        , 'Ihre Rechnung/Referenz Datum.*\n\d* (.*?) '
        , 'Auftrags-Nr.*\n.*\n(.*?) '
        , 'Artikelnummer.*\n.*\n\d* (.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d* (.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d*\s*(.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d*\s*\d*\s*(.*?) '
        , 'Artikelnummer.*\n.*\n.* (.*?)\n'
        , 'netto\s*(.*?)\s*\n'
        , 'USt\s*(.*?%) '
        , 'USt.*%\s*(.*?)\s*\n'
        , 'Gesamt\s*(.*?)\n']

    words = []
    new_name = ['OTTO_Qty Diff']
    for index in range(0,len(rules)):
        rule = rules[index]
        word = re.findall(rule,text)
        if index == 1:
            new_name.append(word)
        if index == 2:
            word = '.'.join(list(reversed(word[0].split('.'))))
            word_list = word.replace('.','')
            new_name.append(word_list)
        if index == 4:
            word = '.'.join(list(reversed(word[0].split('.'))))
        if index >= 8:
            word = word[0].replace('.','').replace(',','.')
        words.append(word)
    for index in range(len(words)):
        if index >= 8:
            if words[index].find('%') >=0 :
                words[index] = float(words[index].replace('%','').strip())/100
            else:
                words[index] = float(words[index])
    return words,new_name


# 价差表解析规则
def pdf_price(text):
    rules= ['.*(B\w* P\w*)\n'
        , 'Unsere Referenz.*\n.*\n.* (.*?)\n'
        , 'Datum\n.*-\S*\s(.*?)\n'
        , 'Ihre Rechnung.*\n(\d*?) '
        , 'Ihre Rechnung/Referenz Datum.*\n\d* (.*?) '
        , 'Auftrags-Nr.*\n.*\n(.*?) '
        , 'Artikelnummer.*\n.*\n\d* (.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d* (.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d*\s*(.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d*\s*\d*\s*(.*?) '
        , 'Artikelnummer.*\n.*\n\d*\s*\d*\s*\d*\s*\d*,\d*\s*(.*?) '
        , 'Artikelnummer.*\n.*\n.* (.*?)\n'
        , 'netto\s*(.*?)\s*\n'
        , 'USt\s*(.*?%) '
        , 'USt.*%\s*(.*?)\s*\n'
        , 'Gesamt\s*(.*?)\n']

    words = []
    new_name = ['OTTO_Price Diff']

    for index in range(0,len(rules)):

        rule = rules[index]
        word = re.findall(rule,text)
        if index == 1:
            new_name.append(word)
        if index == 2:
            word = '.'.join(list(reversed(word[0].split('.'))))
            word_list = word.replace('.','')
            new_name.append(word_list)
        if index == 4:
            word = '.'.join(list(reversed(word[0].split('.'))))
        if index >= 8:
            word = word[0].replace('.','').replace(',','.')
        words.append(word)

    for index in range(len(words)):
        if index >= 8 :
            if words[index].find('%')>=0:
                words[index] = float(words[index].replace('%','').strip())/100
            else:
                words[index] = float(words[index])

    return words,new_name


# OTTO表解析规则
def pdf_otto(text,table,row):
    # print(f'第{row}行数据')
    words = ['OTTO', 'OTTO Payment advice']
    rules1 = rule1 = ['Überweisung Datum.*\n(\d*?) '
                    ,'Überweisung Datum.*\n\d*\s*(.*?)\n'
                    ,'Überweisung Datum.*\n\d*\s*(.*?)\n']
    for index in range(len(rules1)):
        rule1 = rules1[index]
        word = re.findall(rule1, text)
        if (index == 2):
            word = '.'.join(list(reversed(word[0].split('.'))))
        words.append(word)
    for index in range(7):
        if index == 0:
            if bool(re.search(r'\d', table[row][0])):
                if (table[row][0].find('/') >= 0):
                    strs = re.split(r'(\d+/\d+)', table[row][index])
                    words.append(strs[0])
                    words.append(strs[1])
                elif (table[row][0].find(')') >= 0):
                    num = table[row][index].index(')')
                    words.append(table[row][index][0:num+1])
                    words.append(table[row][index][num+1:len(table[row][index])])
                else:
                    strs = re.split(r'(\d+)', table[row][index])
                    words.append(strs[0])
                    words.append(strs[1])
            else :
                words.append(table[row][0])
        elif index == 1:
            if bool(re.search(r' ',table[row][index])):
                strs = table[row][index].split(' ')
                words.append(strs[0])
                words.append(strs[1])
                Bill_date = '.'.join(list(reversed(strs[1].split('.'))))
                words.append(Bill_date)
            else:
                words.append(table[row][index])
                Bill_date ='.'.join(list(reversed(table[row][index].split('.'))))
                words.append(Bill_date)
        elif index > 1:
            if index >= 3 and index != 5:
                if bool(re.search(r'-',table[row][index])):
                    wd = float(('-'+str(table[row][index]).replace('.','').replace(',','.').replace('-','')))
                    words.append(wd)
                    continue
                words.append(float(str(table[row][index]).replace('.','').replace(',','.')))
                continue
            if index == 5:
                words.append(float(table[row][index]))
                continue
            words.append(table[row][index])
    return words


# 退机表名字
def pdf_return_name(texts):
    new_name = ['OTTO_ASS Return']
    rules1 = ['^.*\n.*\n.*\n.*\n.*\n\s*(\d*?)\n'
        ,'^.*\n.*\n.*\n.*\n.*\n.*\n.*\n(.*?) ']
    for index in range(len(rules1)):
        rule1 = rules1[index]
        word = re.findall(rule1, texts[0])
        # print(word)
        if index == 1 :
            word = '20'+''.join(list(reversed(word[0].split('.'))))
        new_name.append(word)
    return new_name


# 退机表解析规则
def pdf_return(texts,tables,row,id):
    # print(f'第{row}行数据')
    words = ['OTTO', 'OTTO DN']
    rules1 = ['^.*\n.*\n.*\n.*\n.*\n\s*(\d*?)\n'
        , '^.*\n.*\n.*\n.*\n.*\n.*\n\s*(.*?)\n'
        , '^.*\n.*\n.*\n.*\n.*\n.*\n.*\n(.*?) '
        , '^.*\n.*\n.*\n.*\n.*\n.*\n.*\n(.*?) '
        , '^.*\n.*\n.*\n.*\n.*\n.*\n.*\n\d*\.(.*?)\.']
    for index in range(len(rules1)):
        rule1 = rules1[index]
        word = re.findall(rule1, texts[0])
        if index == 2:
            word =word[0]+ '20'
        if index == 3 :
            word = '20'+'.'.join(list(reversed(word[0].split('.'))))
        words.append(word)
    if id == 0 :
        for i in range(8):
            if i == 2:
                word =  tables[0][2][i].split('\n')
                words.append(word)
            elif i == 3:
                word =  tables[0][2][i].split('\n')[row]
                words.append(word)
            elif i == 4:
                words.append(1)
            elif i == 6:
                word = table[0][2][5].replace(' ', '').replace(',','.').split('\n')[row]
                words.append(word)
            else:
                word = table[0][2][i].replace(' ', '').replace(',','.').split('\n')[row]
                words.append(word)
    elif id == 1 :
        for i in range(8):
            if i == 2:
                word = tables[id][2][i].split('\n')
                words.append(word)
            elif i == 3:
                word = tables[id][2][i].split('\n')[row+1]
                words.append(word)
            elif i == 4:
                words.append(1)
            elif i == 6:
                word = table[id][2][5].replace(' ', '').replace(',', '.').split('\n')[row]
                words.append(word)
            else:
                word = table[id][2][i].replace(' ', '').replace(',', '.').split('\n')[row]
                words.append(word)
    rules2 = ['PAUSCHALVERGUETUNG\s*(.*?):'
        , 'dadaf(.*?)\n'
        , 'ZWISCHENSUMME:\s*(.*?)\n'
        , 'VORSTEUER:\s*(.*?)\n'
        , 'GESAMTBELASTUNG:\s*(.*?)\n']
    for index in range(len(rules2)):
        rule2 = rules2[index]
        word = re.findall(rule2, texts[len(texts)-1])
        # print(word)
        if index >=2:
            word = word[0].replace('.','').replace(',','.')
        words.append(word)
    for index in range(len(words)):
        if index >= 11 and index != 16:
            if type(words[index])== list :
                if words[index][0].find('%') >=0:
                    words[index] = int(words[index][0].replace('%',''))/100
                else:
                    words[index] = float(words[index][0])
            else:
                if str(words[index]).find('%') >= 0:
                    words[index] = int(words[index].replace('%',''))/100
                else:
                    words[index] = float(words[index])
    return words
    

#存入excel中
def save_table(sheet_name,wk,words):
    if (sheet_name == '年返'):
        sheet1 = wk[sheet_name]
        num = len(words)
        row_cnt = sheet1.max_row + 1
        if num > 13:
            for j in range(0, 13):
                if type(words[j])== float :
                    str =words[j]
                    sheet1.cell(row=row_cnt, column=j + 1, value=str)
                else:
                    str = ''.join(words[j]).strip()
                    sheet1.cell(row=row_cnt, column=j + 1, value=str)
            for j in range(0, 9):
                sheet1.merge_cells(start_row=row_cnt, start_column=j + 1, end_row=row_cnt + 1, end_column=j + 1)
            sheet1.merge_cells(start_row=row_cnt, start_column=13, end_row=row_cnt + 1, end_column=13)

            str10 = words[13]
            str11 = ''.join(words[14]).strip()
            str12 = words[15]
            sheet1.cell(row=row_cnt + 1, column=10, value=str10)
            sheet1.cell(row=row_cnt + 1, column=11, value=str11)
            sheet1.cell(row=row_cnt + 1, column=12, value=str12)
        else:
            for j in range(0, num):
                str = ''.join(words[j]).strip()
                sheet1.cell(row=row_cnt, column=j + 1, value=str)
        return
    table = wk[sheet_name]
    # sheet1 = wk.active
    # print(sheet1.title)
    row_cnt = table.max_row + 1
    # print(row_cnt)
    for index in range(len(words)):
        # print(words[index])
        if isinstance(words[index] ,float):
            # word = ''.join(str(words[index])).strip()
            word = words[index]
        elif isinstance(words[index],int):
            word = words[index]
        else:
            # print(type(words[index]))
            word = ''.join(words[index]).strip()
            if word.find('%') >= 0:
                word = word.replace(' ', '')
        table.cell(row=row_cnt, column=index + 1, value=word)
    if sheet_name == '退机DN':
        row_cnt = table.max_row + 1
        for i in range(2,row_cnt):
            table[f'Q{i}'] = f'=N{i}*P{i}'
 
#根据匹配规则得到有效词
def find_words(text,rules):
    words = []
    for index in range(0, len(rules)):
        rule = rules[index]
        word = re.findall(rule, text)
        words.append(word)
    return words


# 解析pdf,得到文本和表格
def readPdf_DN(file_name):
    #打开pdf
    pdf = pdfplumber.open(pdf_path)
    # 解析当前文件下的pdf
    texts = []
    tables = []
    for page in pdf.pages:
        text = page.extract_text()
        texts.append(text)
        table = page.extract_table()
        tables.append(table)
    pdf.close()
    return texts,tables


# 解析pdf，得到文本和表格
def readPdf(file_name):
    #打开pdf
    pdf = pdfplumber.open(pdf_path)
    if (file_name.find('OTTO Payment')>=0):
        df = tabula.read_pdf(file_name, encoding='gbk', pages='all')
        # print(df)
        for page in pdf.pages:
            text = page.extract_text()
            pdf.close()
            return text,df
    elif (file_name.find('Rework')>=0):
        page = pdf.pages[1]
        text = page.extract_text()
        table = page.extract_table()
        pdf.close()
        return text,table
    elif (file_name.find('年返')>=0):
        page = pdf.pages[0]
        text = page.extract_text()
        table = page.extract_table()
        pdf.close()
        return text,table
    elif (file_name.find('退机DN')>=0):
        texts = []
        tables = []
        for page in pdf.pages:
            # page = pdf.pages[0]
            text = page.extract_text()
            texts.append(text)
            # print(text)
            table = page.extract_table()
            tables.append(table)
            # print(table)
        pdf.close()
        return texts,tables
    # 解析当前文件下的pdf
    # for page in pdf.pages:
    page = pdf.pages[0]
    text = page.extract_text()
    table = page.extract_table()
    pdf.close()
    return text,table


# 创建excel结果表
def create_table(wk):
    # 如果excel存在删除结果表
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)
        except PermissionError:
            print(f'{excel_path},The file in the directory is open. Please close it and execute again')
            os.system("pause")
            sys.exit()
    # 创建table
    wk['Sheet'].title = sheet_names[0]
    for i in range(1,len(sheet_names)):
        wk.create_sheet(sheet_names[i])
    # table表头
    # Mark
    head_M = ['Type', 'Beleg No.', 'Beleg Datum', 'Content', 'Period', 'Bezeichnung', 'Product Type', 'Menge', 'Einzelpreis (netto)', 'Gesamt (netto)', 'MwSt.', 'Ust', 'Gesamt Rechnungsbetrag']
    # Rework
    head_R = ['Beleg No.','KONTROLLBERICHTSNUMMER','LIEFERANTENKENNZIFFER','EK','Saison','LIEFERMENGE','AUFTRAG NR.','ANZ-NR.','ARTIKEL NR.','STYLE NR.','ARTIKELBESCHREIBUNG','REKLAMATIONSGRUND','NACHBEARBEITUNGSANWEISUNG','Leistungsdatum','NACHBEARBEITETE MENGE','ohne Fehler','korrigiert (lagerfähig)','abgelehnt (nicht lagerfähig)','ORIGINAL KOSTEN','Nachbearbeitungskosten / Sonstiges','Warenwert','Reduktionsbetrag','Claim Kosten','MwSt','Ust.','Gesamt']
    # 年返
    head_Y = ['Rechnungsnummer','Belegdatum','Time','Umsatzanteil (HAWA)','Konditionenart','Ausprägung','soll','Bereits realisiert','differenz','MwSt','Kond. rel. Umsatz','Ust','Gesamt Brutto Zahlbetrag*']
    # 量差
    head_L = ['Name','Unsere Referenz','DN datum','Ihre refe','rechnung datum','Auftrags-Nr.','Artikelnummer','Grösse','Menge','Rechnungspreis','Betrag','netto','MwSt','Ust','Gesamt']
    # 价差
    head_P = ['Name','Unsere Referenz','DN datum','Ihre Referenz','Rechnung datum','Auftrags-Nr.','Artikelnummer','Grösse','Menge','Rechnungspreis','Auftragspreis','Differenz','netto','MwSt','Ust','Gesamt']
    # OTTO
    head_V = ['Channel','Name','Überweisung','Datum','Repay_date','Buchungstext','Belegnummer','Beleg-datum','Bill_date','WLS','Betrag','Quellen-steuer','%','Skonto']
    # 退机
    head_N = ['Channel','Name','Belegnummer','Rechnung-Nr','Datum','Date','Month','POS','ART-NR','GRS','ARTIKELBEZEICHNUNG','MENGE','EINZEL-PREIS','GESAMT-PREIS','MWST','PAUSCHALVERGUETUNG 0%/15%','Calculation','ZWISCHENSUMME','VORSTEUER','GESAMTBELASTUNG']
    heads = [head_V,head_N,head_M,head_R,head_P,head_L,head_Y]
    # print(len(head_M),len(head_R),len(head_Y),len(head_L),len(head_P),len(head_V),len(head_N))
    # 将表头写入表中
    for i in range(0,len(sheet_names)):
        table = wk[sheet_names[i]]
        headWrite(table, heads[i])
    # 保存excel
    wk.save(excel_path)


# 写入表头
def headWrite(sht,head):
    lenths = len(head)
    for lenth in range(0, lenths):
        sht.cell(row=1, column=lenth + 1, value=head[lenth])

if __name__ == '__main__':
    # try:
        # 加载配置文件
        path_dict = loadSettingFile('./KEYWORDS.txt')
        excel_path = 'Result.xlsx'
        FOLDER_RULE = path_dict['FOLDER_RULE'].split(';')
        sheet_names = FOLDER_RULE
        wk = workbook.Workbook()
        create_table(wk)
        # 读取文件夹
        for i in range(0,len(sheet_names)):
            path = os.path.join(os.getcwd(),sheet_names[i])
            if os.path.exists(path):
                print(f'-------Read path【{path}】----------')
                # 获取目录下的全部文件名
                file_names = os.listdir(path)
                # 循环全部文件
                for file_name in range(0, len(file_names)):
                    pdf_path = os.path.join(path,file_names[file_name])
                    print(f'-------read file[{pdf_path}]----------')
                    # 得到pdf的文本和表格
                    text,table = readPdf(pdf_path)
                    # 对对应的目录进行操作
                    if sheet_names[i] == 'Marketing' :
                        # 获取需要的文本、获取重新的命名
                        words,new_name = pdf_marketing(text)
                        # 对文件重命名
                        rename(pdf_path,path,new_name)
                        # print(words)
                        # 数据保存到excel
                        save_table(sheet_names[i],wk,words)
                    elif sheet_names[i] == 'Rework' :
                        # 获取需要的文本
                        words = pdf_Rework(text)
                        # 获取需要的重命名
                        new_name = pdf_Rework_name(text)
                        # pdf重命名
                        rename(pdf_path,path,new_name)
                        # print(words)
                        # 保存excel
                        save_table(sheet_names[i],wk,words)
                    elif sheet_names[i] == '年返' :
                        # 获取新的名字和需要的文本
                        words,new_name = pdf_year(text)
                        # print(words)
                        # 对文件重命名
                        rename(pdf_path, path, new_name)
                        # print('rename')
                        # 保存到excel
                        save_table(sheet_names[i],wk,words)
                    elif sheet_names[i] == '量差':
                        # 获取需要的文本和新的名字
                        words,new_name = pdf_amount(text)
                        # 对pdf重命名
                        rename(pdf_path, path, new_name)
                        # 保存内容到excel
                        save_table(sheet_names[i],wk,words)
                    elif sheet_names[i] == '价差':
                        # 获取需要的文本和新的名字
                        words,new_name = pdf_price(text)
                        # 对pdf重命名
                        rename(pdf_path, path, new_name)
                        # 保存数据到excel
                        save_table(sheet_names[i],wk,words)
                    elif sheet_names[i] == 'OTTO Payment Advice':
                        # 循环处理每一页的表格
                        for table_index in range(len(table)):
                            # 去掉空行
                            table_columns = table[table_index].columns
                            # 删除第6、第7行为空的行的数据
                            table_no = table[table_index].loc[table[table_index][table_columns[6]].notnull(), :]
                            table_no = table[table_index].loc[table[table_index][table_columns[5]].notnull(), :]
                            # 对第1为空的替换为‘’
                            table_no = table_no.fillna({table_columns[0]: ''})
                            # table_no = table[table_index].dropna(how='any')
                            # 将 DataFrame 的数据结构转化为array
                            data_array_page = np.array(table_no)
                            # 将array数据结构转化为list
                            data_list_page = data_array_page.tolist()
                            # 对pdf的第一页的表格做特殊处理，将第1行的数据处理后存入list
                            if table_index == 0:
                                # print('----',table_columns)
                                if table_columns[0] == 'Unnamed: 0':
                                    table_columns_list = []
                                    count = 0
                                    for ii in table_columns:
                                        count += 1
                                        if ii == 'Unnamed: 0':
                                            table_columns_list.append('')
                                            continue
                                        if ii == '0,00.1':
                                            table_columns_list.append('0,00')
                                            continue
                                        table_columns_list.append((ii))
                                    # print(table_columns_list)
                                    data_list_page.insert(0, table_columns_list)
                                else:
                                    # print(table_columns)
                                    data_list_page.insert(0, table_columns)
                            # 循环每一行的数据然后存入excel
                            for row in range(len(data_list_page)):
                                words = pdf_otto(text, data_list_page, row)
                                # print(words)
                                save_table(sheet_names[i],wk,words)
                    elif sheet_names[i] == '退机DN':
                        texts, tables = text,table
                        # 对特殊逻辑的退机表格不做处理
                        if tables[0] == None :
                            continue
                        else:
                            # 得到新的名字
                            new_name = pdf_return_name(texts)
                            # 对pdf重命名
                            rename(pdf_path, path, new_name)
                            # 对每一页的pdf做处理
                            for index in range(len(tables)):
                                # 得到每一页pdf有多少行
                                rows = len(table[index][2][0].replace(' ', '').split('\n'))
                                # 对pdf每行的数据做处理
                                for row in range(rows):
                                    menges = table[index][2][4].replace(' ', '').split('\n')
                                    # 得到每条退机记录有多少台
                                    menge = menges[row]
                                    # 每条记录有多少台,循环多少次
                                    for number in range(int(menge)):
                                        words = pdf_return(texts, tables, row, index)
                                        save_table(sheet_names[i], wk, words)
        wk.save(excel_path)
        wk.close()
        os.system("pause")
        print('The program is running and exiting normally。。。')
        time.sleep(2)
    # except Exception as e :
    #     print('The program execution is abnormal')
    #     print('Unknown exception',e,type(e),'Please remove the error！！！')
    #     os.system("pause")
    #     time.sleep(3)























# for table in page.extract_tables():
    # 得到的table是嵌套list类型，转化成DataFrame更加方便查看和分析
    # df = pd.DataFrame(table[1:], columns=table[0])
    # df.to_excel('result.xlsx')

# camelot 方法解析不完全
# tables = camelot.read_pdf(filepath=path,pages='1',flavor='stream')
# df = pd.DataFrame(tables[0].data)
# print(df)

# pyPDF2 方法解析不完全
# mypdf = open(path,mode='rb')
# pdf_document = PyPDF2.PdfFileReader(mypdf)
# pdf_document.numPages
# first_page = pdf_document.getPage(0)
# print(first_page.extractText())
